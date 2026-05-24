from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


TARGET_SHEET = "経営計画"
SCHEDULE_SHEET = "栽培スケジュール"
ASSET_SHEET = "固定資産"

MONTH_START_COL = 3  # C
MONTH_COUNT = 49  # C:AY
ANNUAL_MONTHS = 48  # C:AX, four fiscal years
FIELDS = ("A", "B", "C")


@dataclass(frozen=True)
class CropSchedule:
    crop_id: str
    base_crop_id: str
    name: str
    family: str | None
    fallow_years: int
    labor_s: float
    labor_o: float
    labor_f: float
    income: float
    total_labor: float
    gross_revenue: float
    operating_cost: float
    operating_cost_per_month: float
    gross_per_month: float
    depreciation: float | None
    monthly_marks: tuple[str | None, ...]


@dataclass(frozen=True)
class Assets:
    asset_purchase: float
    annual_depreciation: float
    annual_rent_thousand_yen: float
    land_purchase_thousand_yen: float
    plot_count: float | None = None


@dataclass(frozen=True)
class CurrentPlan:
    month_labels: list[int]
    crop_ids: dict[str, list[str | None]]
    marks: dict[str, list[str | None]]


@dataclass(frozen=True)
class CalculationResult:
    crop_names: dict[str, list[str | None]]
    labor_by_field: dict[str, list[float]]
    labor_total: list[float]
    revenue_by_field: dict[str, list[float]]
    cost_by_field: dict[str, list[float]]
    revenue_total: list[float]
    cost_total: list[float]
    monthly_income: list[float]
    leasing: dict[str, list[float]]
    purchase: dict[str, list[float]]


def norm_id(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def num(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)


def read_crop_schedules(path: Path) -> dict[str, CropSchedule]:
    workbook = load_workbook(path, data_only=True)
    ws = workbook[SCHEDULE_SHEET]
    crops: dict[str, CropSchedule] = {}
    headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}

    id_col = header_col(headers, ("#作付", "#"), 1)
    base_id_col = header_col(headers, ("#品目", "参照用#"), 2)
    name_col = header_col(headers, "野菜名", 3)
    income_col = header_col(headers, "所得（千円）", 10)
    gross_col = header_col(headers, "粗収益", 24)
    cost_col = header_col(headers, "経営費(減価償却費除く)", 25)
    cost_per_month_col = headers.get("経営費/月")
    total_labor_col = header_col(headers, "総労働時間(h)", 11)
    fallow_col = headers.get("休栽年数")
    labor_s_col = header_col(headers, "s(h)", 7)
    labor_o_col = header_col(headers, "o(h)", 8)
    labor_f_col = header_col(headers, "f(h)", 9)
    gross_per_month_col = header_col(headers, "粗収益/月", 27)
    depreciation_col = headers.get("減価償却費")
    month_cols_by_label = [headers.get(label) for label in [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]]
    if any(col is None for col in month_cols_by_label):
        month_cols_by_label = list(range(12, 24))

    for row in range(2, ws.max_row + 1):
        crop_id = norm_id(ws.cell(row, id_col).value)
        if crop_id is None:
            continue
        base_crop_id = norm_id(ws.cell(row, base_id_col).value) or crop_id
        crops[crop_id] = CropSchedule(
            crop_id=crop_id,
            base_crop_id=base_crop_id,
            name=str(ws.cell(row, name_col).value),
            family=None,
            fallow_years=int(num(ws.cell(row, fallow_col).value)) if fallow_col else 0,
            labor_s=num(ws.cell(row, labor_s_col).value),
            labor_o=num(ws.cell(row, labor_o_col).value),
            labor_f=num(ws.cell(row, labor_f_col).value),
            income=num(ws.cell(row, income_col).value),
            total_labor=num(ws.cell(row, total_labor_col).value),
            gross_revenue=num(ws.cell(row, gross_col).value),
            operating_cost=num(ws.cell(row, cost_col).value),
            operating_cost_per_month=(
                num(ws.cell(row, cost_per_month_col).value)
                if cost_per_month_col
                else num(ws.cell(row, cost_col).value)
            ),
            gross_per_month=num(ws.cell(row, gross_per_month_col).value),
            depreciation=ws.cell(row, depreciation_col).value if depreciation_col else None,
            monthly_marks=tuple(
                norm_mark(ws.cell(row, col).value) for col in month_cols_by_label
            ),
        )

    if fallow_col is None and "経営指標" in workbook.sheetnames:
        indicator = workbook["経営指標"]
        indicators: dict[str, tuple[str | None, int]] = {}
        for row in range(2, indicator.max_row + 1):
            for col in (1, 2):
                crop_id = norm_id(indicator.cell(row, col).value)
                if crop_id is not None:
                    indicators[crop_id] = (
                        indicator.cell(row, 7).value,
                        int(num(indicator.cell(row, 8).value)),
                    )

        for crop in crops.values():
            if crop.base_crop_id in indicators:
                family, fallow_years = indicators[crop.base_crop_id]
                object.__setattr__(crop, "family", family)
                object.__setattr__(crop, "fallow_years", fallow_years)

    return crops


def header_col(headers: dict[Any, int], names: Any | tuple[Any, ...], fallback: int) -> int:
    for name in names if isinstance(names, tuple) else (names,):
        if name in headers:
            return headers[name]
    return fallback


def norm_mark(value: Any) -> str | None:
    if value is None or value == "":
        return None
    return str(value)


def read_assets(path: Path) -> Assets:
    workbook = load_workbook(path, data_only=True)
    ws = workbook[ASSET_SHEET]
    key_values = read_asset_key_values(ws)
    if key_values:
        return Assets(
            asset_purchase=key_values["asset_purchase"],
            annual_depreciation=key_values["annual_depreciation"],
            annual_rent_thousand_yen=key_values["annual_rent_thousand_yen"],
            land_purchase_thousand_yen=key_values["land_purchase_thousand_yen"],
            plot_count=key_values.get("plot_count"),
        )
    if ws["B17"].value == "区画数":
        plot_count = num(ws["B18"].value)
        return Assets(
            asset_purchase=num(ws["D14"].value),
            annual_depreciation=num(ws["H14"].value),
            annual_rent_thousand_yen=num(ws["F21"].value) / 1000,
            land_purchase_thousand_yen=num(ws["F24"].value) / 1000,
            plot_count=plot_count,
        )
    return Assets(
        asset_purchase=num(ws["D14"].value),
        annual_depreciation=num(ws["H14"].value),
        annual_rent_thousand_yen=num(ws["B18"].value) / 1000,
        land_purchase_thousand_yen=4431.0,
    )


def read_asset_key_values(ws) -> dict[str, float]:
    required = {
        "asset_purchase",
        "annual_depreciation",
        "annual_rent_thousand_yen",
        "land_purchase_thousand_yen",
    }
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column):
            if ws.cell(row, col).value == "キー" and ws.cell(row, col + 1).value == "値":
                values: dict[str, float] = {}
                current = row + 1
                while current <= ws.max_row:
                    key = ws.cell(current, col).value
                    if key is None or key == "":
                        break
                    values[str(key)] = num(ws.cell(current, col + 1).value)
                    current += 1
                return values if required.issubset(values) else {}
    return {}


def assets_for_field_count(assets: Assets, field_count: int) -> Assets:
    if not assets.plot_count or assets.plot_count == field_count:
        return assets
    scale = field_count / assets.plot_count
    return Assets(
        asset_purchase=assets.asset_purchase,
        annual_depreciation=assets.annual_depreciation,
        annual_rent_thousand_yen=assets.annual_rent_thousand_yen * scale,
        land_purchase_thousand_yen=assets.land_purchase_thousand_yen * scale,
        plot_count=float(field_count),
    )


def read_current_plan(path: Path) -> CurrentPlan:
    workbook = load_workbook(path, data_only=True)
    ws = workbook[TARGET_SHEET]
    month_labels = [int(ws.cell(2, col).value) for col in month_cols()]

    crop_ids: dict[str, list[str | None]] = {}
    marks: dict[str, list[str | None]] = {}
    for field, crop_row, mark_row in zip(FIELDS, (3, 4, 5), (9, 10, 11)):
        crop_ids[field] = [norm_id(ws.cell(crop_row, col).value) for col in month_cols()]
        marks[field] = [norm_mark(ws.cell(mark_row, col).value) for col in month_cols()]

    return CurrentPlan(month_labels=month_labels, crop_ids=crop_ids, marks=marks)


def month_cols() -> range:
    return range(MONTH_START_COL, MONTH_START_COL + MONTH_COUNT)


def labor_for_mark(crop: CropSchedule, mark: str | None) -> float:
    if mark is None:
        return 0.0
    if mark == "s":
        return crop.labor_s
    if mark == "o":
        return crop.labor_o
    return crop.labor_f


def calculate(plan: CurrentPlan, crops: dict[str, CropSchedule], assets: Assets) -> CalculationResult:
    crop_names = {field: zero_texts() for field in FIELDS}
    labor_by_field = {field: zeroes() for field in FIELDS}
    revenue_by_field = {field: zeroes() for field in FIELDS}
    cost_by_field = {field: zeroes() for field in FIELDS}

    for field in FIELDS:
        for month in range(MONTH_COUNT):
            crop_id = plan.crop_ids[field][month]
            mark = plan.marks[field][month]
            if crop_id is None or mark is None:
                continue
            crop = crops[crop_id]
            crop_names[field][month] = crop.name
            labor_by_field[field][month] = labor_for_mark(crop, mark)
            if mark == "f":
                revenue_by_field[field][month] = crop.gross_per_month
            if mark == "s":
                cost_by_field[field][month] = crop.operating_cost_per_month

    labor_total = sum_fields(labor_by_field)
    revenue_total = sum_fields(revenue_by_field)
    cost_total = sum_fields(cost_by_field)
    monthly_income = [revenue_total[i] - cost_total[i] for i in range(MONTH_COUNT)]

    leasing = annual_summaries(
        revenue_total,
        cost_total,
        assets,
        add_rent=True,
        initial_cash=10000.0,
        land_purchase=0.0,
    )
    purchase = annual_summaries(
        revenue_total,
        cost_total,
        assets,
        add_rent=False,
        initial_cash=15000.0,
        land_purchase=assets.land_purchase_thousand_yen,
    )

    return CalculationResult(
        crop_names=crop_names,
        labor_by_field=labor_by_field,
        labor_total=labor_total,
        revenue_by_field=revenue_by_field,
        cost_by_field=cost_by_field,
        revenue_total=revenue_total,
        cost_total=cost_total,
        monthly_income=monthly_income,
        leasing=leasing,
        purchase=purchase,
    )


def zeroes() -> list[float]:
    return [0.0 for _ in range(MONTH_COUNT)]


def zero_texts() -> list[str | None]:
    return [None for _ in range(MONTH_COUNT)]


def sum_fields(values: dict[str, list[float]]) -> list[float]:
    return [sum(values[field][i] for field in FIELDS) for i in range(MONTH_COUNT)]


def annual_blocks(values: list[float]) -> list[float]:
    return [sum(values[start : start + 12]) for start in range(0, ANNUAL_MONTHS, 12)]


def annual_summaries(
    revenue_total: list[float],
    cost_total: list[float],
    assets: Assets,
    *,
    add_rent: bool,
    initial_cash: float,
    land_purchase: float,
) -> dict[str, list[float]]:
    gross = annual_blocks(revenue_total)
    operating_cost = annual_blocks(cost_total)
    if add_rent:
        operating_cost = [value + assets.annual_rent_thousand_yen for value in operating_cost]

    depreciation = [assets.annual_depreciation for _ in gross]
    income = [gross[i] - operating_cost[i] - depreciation[i] for i in range(4)]

    trial_gross = [gross[0] * 0.7, gross[1] * 0.9, gross[2], gross[3]]
    trial_cost = operating_cost[:]
    trial_depreciation = depreciation[:]
    trial_income = [
        trial_gross[i] - trial_cost[i] - trial_depreciation[i] for i in range(4)
    ]

    cash_in = trial_gross[:]
    cash_out = trial_cost[:]
    cash_out[0] += assets.asset_purchase + land_purchase
    cash_flow = [cash_in[i] - cash_out[i] for i in range(4)]
    ending_cash: list[float] = []
    balance = initial_cash
    for flow in cash_flow:
        balance += flow
        ending_cash.append(balance)

    return {
        "gross": gross,
        "operating_cost": operating_cost,
        "depreciation": depreciation,
        "income": income,
        "trial_gross": trial_gross,
        "trial_cost": trial_cost,
        "trial_depreciation": trial_depreciation,
        "trial_income": trial_income,
        "cash_in": cash_in,
        "cash_out": cash_out,
        "cash_flow": cash_flow,
        "ending_cash": ending_cash,
    }


def compare_with_workbook(path: Path, result: CalculationResult, tolerance: float = 1e-6) -> list[str]:
    workbook = load_workbook(path, data_only=True)
    ws = workbook[TARGET_SHEET]
    mismatches: list[str] = []

    row_vectors: list[tuple[str, int, Iterable[float]]] = [
        ("labor_total", 18, result.labor_total),
        ("revenue_total", 22, result.revenue_total),
        ("cost_total", 26, result.cost_total),
        ("monthly_income", 30, result.monthly_income),
    ]
    for name, row, values in row_vectors:
        for offset, value in enumerate(values):
            col = MONTH_START_COL + offset
            check_number(ws.cell(row, col).value, value, f"{name} {cell_ref(row, col)}", mismatches, tolerance)

    summary_rows = {
        "leasing": {
            "gross": 35,
            "operating_cost": 36,
            "depreciation": 37,
            "income": 38,
            "trial_gross": 42,
            "trial_cost": 43,
            "trial_depreciation": 44,
            "trial_income": 45,
            "cash_in": 49,
            "cash_out": 50,
            "cash_flow": 51,
            "ending_cash": 52,
        },
        "purchase": {
            "gross": 57,
            "operating_cost": 58,
            "depreciation": 59,
            "income": 60,
            "trial_gross": 64,
            "trial_cost": 65,
            "trial_depreciation": 66,
            "trial_income": 67,
            "cash_in": 71,
            "cash_out": 72,
            "cash_flow": 73,
            "ending_cash": 74,
        },
    }

    for scenario, rows in summary_rows.items():
        data = getattr(result, scenario)
        for key, row in rows.items():
            for i, value in enumerate(data[key]):
                col = MONTH_START_COL + i
                check_number(ws.cell(row, col).value, value, f"{scenario}.{key} {cell_ref(row, col)}", mismatches, tolerance)

    return mismatches


def check_number(actual: Any, expected: float, label: str, mismatches: list[str], tolerance: float) -> None:
    actual_num = num(actual)
    if abs(actual_num - expected) > tolerance:
        mismatches.append(f"{label}: workbook={actual_num:.10g}, python={expected:.10g}")


def cell_ref(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def check_constraints(plan: CurrentPlan, crops: dict[str, CropSchedule]) -> list[str]:
    issues: list[str] = []
    for field in FIELDS:
        harvest_months = [
            i for i, mark in enumerate(plan.marks[field]) if mark is not None and "f" in mark
        ]
        for month in harvest_months:
            for next_month in range(month + 1, min(month + 3, MONTH_COUNT)):
                if plan.marks[field][next_month] == "s":
                    issues.append(
                        f"{field}: harvest at {month_label(month)} but next sowing at {month_label(next_month)}"
                    )

        last_harvest_by_crop: dict[str, tuple[int, str]] = {}
        for month, (crop_id, mark) in enumerate(zip(plan.crop_ids[field], plan.marks[field])):
            if crop_id is None or mark is None:
                continue
            crop = crops[crop_id]
            base_crop_id = crop.base_crop_id
            if mark == "s" and base_crop_id in last_harvest_by_crop:
                last_harvest_month, last_crop_id = last_harvest_by_crop[base_crop_id]
                fallow_months = crop.fallow_years * 12
                elapsed = month - last_harvest_month
                if elapsed < fallow_months:
                    issues.append(
                        f"{field}: {crop_id} starts at {month_label(month)} after {elapsed} months from {last_crop_id}; needs {fallow_months}"
                    )
            if "f" in mark:
                last_harvest_by_crop[base_crop_id] = (month, crop_id)

    return issues


def month_label(index: int) -> str:
    fiscal_year = index // 12
    month = [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3][index % 12]
    return f"FY{fiscal_year} {month}月"
