from __future__ import annotations

import csv
import argparse
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from ortools.sat.python import cp_model

from keiei_plan import (
    ANNUAL_MONTHS,
    FIELDS,
    MONTH_COUNT,
    MONTH_START_COL,
    assets_for_field_count,
    annual_summaries,
    month_label,
    read_assets,
    read_crop_schedules,
)


DEFAULT_WORKBOOK = Path("経営計画_四本20260520.xlsx")
OUT_DIR = Path("out")
OUTPUT_WORKBOOK = OUT_DIR / "経営計画_optimized.xlsx"
SCALE = 1000
FIELD_NAMES = tuple("ABCDEFGHIJKLMNOPQRSTUVWXYZ")


@dataclass(frozen=True)
class Candidate:
    index: int
    field: str
    crop_id: str
    crop_name: str
    fiscal_year: int
    active_months: tuple[int, ...]
    blocked_months: tuple[int, ...]
    first_active: int
    last_harvest: int
    labor: tuple[float, ...]
    revenue: tuple[float, ...]
    cost: tuple[float, ...]
    marks: tuple[str | None, ...]
    profit: float
    fallow_months: int


def main() -> int:
    args = parse_args()
    fields = FIELD_NAMES[: args.fields]
    path = args.workbook
    crops = read_crop_schedules(path)
    assets = assets_for_field_count(read_assets(path), args.fields)
    labor_cap = args.labor_cap
    soft_labor = min(args.soft_labor, labor_cap)
    run_dir = make_run_dir(args.workbook, args.fields, labor_cap, soft_labor)
    candidates = build_candidates(crops, fields)
    print(f"Candidates: {len(candidates)}")

    model = cp_model.CpModel()
    x = [model.NewBoolVar(f"x_{candidate.index}") for candidate in candidates]

    # One field cannot hold overlapping crops, including two blank months after harvest.
    for field in fields:
        for month in range(ANNUAL_MONTHS):
            vars_in_month = [
                x[c.index]
                for c in candidates
                if c.field == field and month in c.blocked_months
            ]
            if vars_in_month:
                model.Add(sum(vars_in_month) <= 1)

    # Same crop in the same field must respect its fallow period.
    by_field_crop: dict[tuple[str, str], list[Candidate]] = {}
    for candidate in candidates:
        by_field_crop.setdefault((candidate.field, candidate.crop_id), []).append(candidate)
    for group in by_field_crop.values():
        for i, first in enumerate(group):
            for second in group[i + 1 :]:
                earlier, later = sorted((first, second), key=lambda c: c.first_active)
                if later.first_active - earlier.last_harvest < earlier.fallow_months:
                    model.Add(x[earlier.index] + x[later.index] <= 1)

    monthly_labor_exprs = []
    over_200 = []
    for month in range(ANNUAL_MONTHS):
        labor_expr = sum(
            int(round(candidate.labor[month] * SCALE)) * x[candidate.index]
            for candidate in candidates
        )
        monthly_labor_exprs.append(labor_expr)
        model.Add(labor_expr <= int(round(labor_cap * SCALE)))

        excess_max = max(0, int(round((labor_cap - soft_labor) * SCALE)))
        excess = model.NewIntVar(0, excess_max, f"over_soft_labor_{month}")
        model.Add(excess >= labor_expr - int(round(soft_labor * SCALE)))
        model.Add(excess >= 0)
        over_200.append(excess)

    profit_expr = sum(
        int(round(candidate.profit * SCALE)) * x[candidate.index]
        for candidate in candidates
    )
    overtime_penalty = 8 * sum(over_200)
    model.Maximize(profit_expr - overtime_penalty)

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 20
    solver.parameters.num_search_workers = 8
    status = solver.Solve(model)
    print(f"Status: {solver.StatusName(status)}")
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return 1

    selected = [candidate for candidate in candidates if solver.BooleanValue(x[candidate.index])]
    selected.sort(key=lambda c: (c.field, c.first_active, c.crop_id))
    print(f"Selected: {len(selected)}")
    print(f"Objective: {solver.ObjectiveValue() / SCALE:.3f}")
    print(f"Operating profit: {sum(c.profit for c in selected):.3f} thousand yen")

    labor = [
        sum(candidate.labor[month] for candidate in selected)
        for month in range(ANNUAL_MONTHS)
    ]
    revenue = [
        sum(candidate.revenue[month] for candidate in selected)
        for month in range(ANNUAL_MONTHS)
    ]
    cost = [
        sum(candidate.cost[month] for candidate in selected)
        for month in range(ANNUAL_MONTHS)
    ]
    leasing = annual_summaries(
        revenue + [0.0],
        cost + [0.0],
        assets,
        add_rent=True,
        initial_cash=10000.0,
        land_purchase=0.0,
    )
    purchase = annual_summaries(
        revenue + [0.0],
        cost + [0.0],
        assets,
        add_rent=False,
        initial_cash=15000.0,
        land_purchase=assets.land_purchase_thousand_yen,
    )
    print(f"Max monthly labor: {max(labor):.2f} h")
    print(f"Months over {soft_labor:g}h: {sum(1 for value in labor if value > soft_labor)}")
    print(
        "Leasing trial avg income FY1-FY3: "
        f"{sum(leasing['trial_income'][1:4]) / 3:.3f} thousand yen"
    )
    print(
        "Purchase trial avg income FY1-FY3: "
        f"{sum(purchase['trial_income'][1:4]) / 3:.3f} thousand yen"
    )

    output_prefix = make_output_prefix(args.workbook, args.fields, labor_cap, soft_labor)
    write_outputs(run_dir, output_prefix, selected, labor, revenue, cost, fields)
    print(f"Wrote: {run_dir / f'{output_prefix}_plan.csv'}")
    print(f"Wrote: {run_dir / f'{output_prefix}_monthly.csv'}")
    if args.fields in (len(FIELDS), 4):
        output_workbook = run_dir / f"{output_prefix}.xlsx"
        saved_path = write_optimized_workbook(
            path, selected, leasing, purchase, fields, output_workbook
        )
        print(f"Wrote: {saved_path}")
    else:
        output_workbook = run_dir / f"{output_prefix}_rotation_labor.xlsx"
        saved_path = write_rotation_labor_workbook(
            path, selected, fields, output_workbook
        )
        print(f"Wrote: {saved_path}")

    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help="Workbook to read.",
    )
    parser.add_argument(
        "--fields",
        type=int,
        default=len(FIELDS),
        help="Number of field blocks to optimize. Excel writeback supports 3 or 4.",
    )
    parser.add_argument(
        "--labor-cap",
        type=float,
        default=240.0,
        help="Hard monthly labor-hour cap.",
    )
    parser.add_argument(
        "--soft-labor",
        type=float,
        default=200.0,
        help="Soft monthly labor-hour threshold used in the objective penalty.",
    )
    args = parser.parse_args()
    if args.fields < 1 or args.fields > len(FIELD_NAMES):
        parser.error("--fields must be between 1 and 26")
    return args


def make_output_prefix(
    workbook_path: Path, field_count: int, labor_cap: float, soft_labor: float
) -> str:
    prefix = f"{workbook_path.stem}_optimized_{field_count}fields"
    if labor_cap != 240.0 or soft_labor != 200.0:
        prefix += f"_cap{format_number(labor_cap)}_soft{format_number(soft_labor)}"
    return prefix


def make_run_dir(
    workbook_path: Path, field_count: int, labor_cap: float, soft_labor: float
) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = f"{field_count}fields"
    if labor_cap != 240.0 or soft_labor != 200.0:
        suffix += f"_cap{format_number(labor_cap)}_soft{format_number(soft_labor)}"
    run_dir = OUT_DIR / f"{workbook_path.stem}_{suffix}_{timestamp}"
    run_dir.mkdir(parents=True, exist_ok=False)
    return run_dir


def format_number(value: float) -> str:
    return str(int(value)) if value.is_integer() else str(value).replace(".", "p")


def build_candidates(crops, fields: tuple[str, ...]) -> list[Candidate]:
    candidates: list[Candidate] = []
    for field in fields:
        for crop in crops.values():
            sowing_offsets = [
                offset
                for offset, mark in enumerate(crop.monthly_marks)
                if mark is not None and "s" in mark
            ]
            if not sowing_offsets:
                continue
            sowing_offset = sowing_offsets[0]

            for fiscal_year in range(4):
                base = fiscal_year * 12
                active_months: list[int] = []
                labor = [0.0 for _ in range(ANNUAL_MONTHS)]
                revenue = [0.0 for _ in range(ANNUAL_MONTHS)]
                cost = [0.0 for _ in range(ANNUAL_MONTHS)]
                marks = [None for _ in range(ANNUAL_MONTHS)]
                harvest_months: list[int] = []

                for cycle_index in range(12):
                    offset = (sowing_offset + cycle_index) % 12
                    year_add = 1 if offset < sowing_offset else 0
                    month = base + year_add * 12 + offset
                    mark = crop.monthly_marks[offset]
                    if month >= ANNUAL_MONTHS or mark is None:
                        continue
                    active_months.append(month)
                    marks[month] = mark
                    if mark == "s":
                        labor[month] = crop.labor_s
                        cost[month] = crop.operating_cost
                    elif mark == "o":
                        labor[month] = crop.labor_o
                    else:
                        labor[month] = crop.labor_f
                        revenue[month] = crop.gross_per_month
                        if "f" in mark:
                            harvest_months.append(month)

                if not active_months or not harvest_months:
                    continue
                last_harvest = max(harvest_months)
                blocked = set(active_months)
                for month in (last_harvest + 1, last_harvest + 2):
                    if month < ANNUAL_MONTHS:
                        blocked.add(month)

                candidates.append(
                    Candidate(
                        index=len(candidates),
                        field=field,
                        crop_id=crop.crop_id,
                        crop_name=crop.name,
                        fiscal_year=fiscal_year,
                        active_months=tuple(active_months),
                        blocked_months=tuple(sorted(blocked)),
                        first_active=min(active_months),
                        last_harvest=last_harvest,
                        labor=tuple(labor),
                        revenue=tuple(revenue),
                        cost=tuple(cost),
                        marks=tuple(marks),
                        profit=sum(revenue) - sum(cost),
                        fallow_months=crop.fallow_years * 12,
                    )
                )

    return candidates


def write_outputs(
    output_dir: Path,
    output_prefix: str,
    selected: list[Candidate],
    labor: list[float],
    revenue: list[float],
    cost: list[float],
    fields: tuple[str, ...],
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    with (output_dir / f"{output_prefix}_plan.csv").open("w", newline="", encoding="utf-8-sig") as fp:
        writer = csv.writer(fp)
        writer.writerow(
            [
                "field",
                "crop_id",
                "crop_name",
                "fiscal_year",
                "first_active",
                "last_harvest",
                "profit_thousand_yen",
            ]
        )
        for candidate in selected:
            writer.writerow(
                [
                    candidate.field,
                    candidate.crop_id,
                    candidate.crop_name,
                    candidate.fiscal_year,
                    month_label(candidate.first_active),
                    month_label(candidate.last_harvest),
                    round(candidate.profit, 3),
                ]
            )

    with (output_dir / f"{output_prefix}_monthly.csv").open("w", newline="", encoding="utf-8-sig") as fp:
        writer = csv.writer(fp)
        writer.writerow(["month_index", "month", "labor_hours", "revenue", "cost", "income"])
        for index, hours in enumerate(labor):
            writer.writerow(
                [
                    index,
                    month_label(index),
                    round(hours, 3),
                    round(revenue[index], 3),
                    round(cost[index], 3),
                    round(revenue[index] - cost[index], 3),
                ]
            )

    crop_grid, mark_grid = selected_to_grids(selected, fields)
    with (output_dir / f"{output_prefix}_grid.csv").open("w", newline="", encoding="utf-8-sig") as fp:
        writer = csv.writer(fp)
        writer.writerow(["row"] + [month_label(index) for index in range(ANNUAL_MONTHS)])
        for field in fields:
            writer.writerow([f"{field}_crop_id"] + crop_grid[field])
            writer.writerow([f"{field}_mark"] + mark_grid[field])


def selected_to_grids(
    selected: list[Candidate],
    fields: tuple[str, ...],
) -> tuple[dict[str, list[str | None]], dict[str, list[str | None]]]:
    crop_grid = {field: [None for _ in range(ANNUAL_MONTHS)] for field in fields}
    mark_grid = {field: [None for _ in range(ANNUAL_MONTHS)] for field in fields}
    for candidate in selected:
        for month, mark in enumerate(candidate.marks):
            if mark is None:
                continue
            crop_grid[candidate.field][month] = candidate.crop_id
            mark_grid[candidate.field][month] = mark
    return crop_grid, mark_grid


def write_optimized_workbook(
    source_path: Path,
    selected: list[Candidate],
    leasing: dict[str, list[float]],
    purchase: dict[str, list[float]],
    fields: tuple[str, ...],
    output_path: Path,
) -> Path:
    workbook = load_workbook(source_path)
    ws = workbook["経営計画"]
    field_count = len(fields)
    source_is_four_field = is_four_field_plan_sheet(ws)
    if field_count == 4 and not source_is_four_field:
        expand_plan_sheet_to_four_fields(ws)
        source_is_four_field = True
    if field_count in (3, 4) and "固定資産" in workbook.sheetnames:
        workbook["固定資産"]["B18"] = field_count

    layout_fields = FIELD_NAMES[:4] if source_is_four_field else fields
    crop_grid, mark_grid = selected_to_grids(selected, layout_fields)
    layout = plan_layout(len(layout_fields))
    crop_rows = {field: layout["crop_start"] + i for i, field in enumerate(layout_fields)}
    name_rows = {field: layout["name_start"] + i for i, field in enumerate(layout_fields)}
    mark_rows = {field: layout["mark_start"] + i for i, field in enumerate(layout_fields)}
    set_plan_formulas(ws, layout_fields, layout)

    for field in layout_fields:
        for offset in range(MONTH_COUNT):
            col = MONTH_START_COL + offset
            crop_cell = ws.cell(crop_rows[field], col)
            name_cell = ws.cell(name_rows[field], col)
            mark_cell = ws.cell(mark_rows[field], col)
            if offset < ANNUAL_MONTHS:
                crop_cell.value = excel_crop_id(crop_grid[field][offset])
                mark_cell.value = mark_grid[field][offset]
            else:
                crop_cell.value = None
                mark_cell.value = None

    if "探索結果" in workbook.sheetnames:
        del workbook["探索結果"]
    summary = workbook.create_sheet("探索結果")
    write_summary_sheet(summary, selected, leasing, purchase)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output_path.parent.mkdir(exist_ok=True)
    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        fallback = output_path.with_name(f"{output_path.stem}_new{output_path.suffix}")
        workbook.save(fallback)
        return fallback


def write_rotation_labor_workbook(
    source_path: Path,
    selected: list[Candidate],
    fields: tuple[str, ...],
    output_path: Path,
) -> Path:
    workbook = load_workbook(source_path)
    if "経営計画" in workbook.sheetnames:
        del workbook["経営計画"]
    ws = workbook.create_sheet("経営計画")

    crop_grid, mark_grid = selected_to_grids(selected, fields)
    write_rotation_labor_sheet(ws, fields, crop_grid, mark_grid, selected)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        fallback = output_path.with_name(f"{output_path.stem}_new{output_path.suffix}")
        workbook.save(fallback)
        return fallback


def write_rotation_labor_sheet(
    ws,
    fields: tuple[str, ...],
    crop_grid: dict[str, list[str | None]],
    mark_grid: dict[str, list[str | None]],
    selected: list[Candidate],
) -> None:
    month_labels = [month_number(index) for index in range(ANNUAL_MONTHS)]
    labor_by_field = {
        field: [
            sum(candidate.labor[month] for candidate in selected if candidate.field == field)
            for month in range(ANNUAL_MONTHS)
        ]
        for field in fields
    }
    labor_total = [
        sum(labor_by_field[field][month] for field in fields)
        for month in range(ANNUAL_MONTHS)
    ]

    ws["A1"] = "＜輪作計画表＞"
    write_month_header(ws, 2, month_labels)
    crop_start = 3
    name_start = crop_start + len(fields)
    mark_start = name_start + len(fields)
    for i, field in enumerate(fields):
        crop_row = crop_start + i
        name_row = name_start + i
        mark_row = mark_start + i
        ws.cell(crop_row, 2).value = field
        ws.cell(name_row, 2).value = field
        ws.cell(mark_row, 2).value = field
        for month in range(ANNUAL_MONTHS):
            col = MONTH_START_COL + month
            crop_id = crop_grid[field][month]
            ws.cell(crop_row, col).value = excel_crop_id(crop_id)
            ws.cell(name_row, col).value = crop_name_for_month(selected, field, month)
            ws.cell(mark_row, col).value = mark_grid[field][month]

    labor_title_row = mark_start + len(fields) + 2
    ws.cell(labor_title_row, 1).value = "＜労働時間算定表＞"
    ws.cell(labor_title_row, MONTH_START_COL + ANNUAL_MONTHS - 1).value = "単位：時間"
    write_month_header(ws, labor_title_row + 1, month_labels)
    labor_start = labor_title_row + 2
    for i, field in enumerate(fields):
        row = labor_start + i
        ws.cell(row, 2).value = field
        for month in range(ANNUAL_MONTHS):
            ws.cell(row, MONTH_START_COL + month).value = labor_by_field[field][month]
    total_row = labor_start + len(fields)
    ws.cell(total_row, 2).value = "合計"
    for month in range(ANNUAL_MONTHS):
        ws.cell(total_row, MONTH_START_COL + month).value = labor_total[month]

    summary_row = total_row + 2
    ws.cell(summary_row, 1).value = "最大月労働時間"
    ws.cell(summary_row, 2).value = max(labor_total) if labor_total else 0
    ws.cell(summary_row + 1, 1).value = "200時間超過月数"
    ws.cell(summary_row + 1, 2).value = sum(1 for value in labor_total if value > 200)

    for col in range(1, MONTH_START_COL + ANNUAL_MONTHS):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 10
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 8


def write_month_header(ws, row: int, month_labels: list[int]) -> None:
    for month, label in enumerate(month_labels):
        ws.cell(row, MONTH_START_COL + month).value = label


def month_number(index: int) -> int:
    return [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3][index % 12]


def crop_name_for_month(
    selected: list[Candidate], field: str, month: int
) -> str | None:
    for candidate in selected:
        if candidate.field == field and candidate.marks[month] is not None:
            return candidate.crop_name
    return None


def expand_plan_sheet_to_four_fields(ws) -> None:
    for row in (6, 10, 14, 21, 30, 35):
        ws.insert_rows(row)
        copy_row_style(ws, row - 1, row)


def is_four_field_plan_sheet(ws) -> bool:
    return [ws.cell(row, 2).value for row in range(3, 15)] == [
        "A",
        "B",
        "C",
        "D",
        "A",
        "B",
        "C",
        "D",
        "A",
        "B",
        "C",
        "D",
    ]


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format


def plan_layout(field_count: int) -> dict[str, int]:
    crop_start = 3
    name_start = crop_start + field_count
    mark_start = name_start + field_count
    labor_title = mark_start + field_count + 1
    labor_month = labor_title + 1
    labor_start = labor_month + 1
    labor_total = labor_start + field_count
    pl_title = labor_total + 2
    pl_month = pl_title + 1
    revenue_total = pl_month + 1
    revenue_start = revenue_total + 1
    cost_total = revenue_start + field_count
    cost_start = cost_total + 1
    income = cost_start + field_count
    summary_start = income + 2
    return {
        "crop_start": crop_start,
        "name_start": name_start,
        "mark_start": mark_start,
        "labor_title": labor_title,
        "labor_month": labor_month,
        "labor_start": labor_start,
        "labor_total": labor_total,
        "pl_title": pl_title,
        "pl_month": pl_month,
        "revenue_total": revenue_total,
        "revenue_start": revenue_start,
        "cost_total": cost_total,
        "cost_start": cost_start,
        "income": income,
        "summary_start": summary_start,
    }


def set_plan_formulas(ws, fields: tuple[str, ...], layout: dict[str, int]) -> None:
    for i, field in enumerate(fields):
        ws.cell(layout["crop_start"] + i, 2).value = field
        ws.cell(layout["name_start"] + i, 2).value = field
        ws.cell(layout["mark_start"] + i, 2).value = field
        ws.cell(layout["labor_start"] + i, 2).value = field
        ws.cell(layout["revenue_start"] + i, 2).value = field
        ws.cell(layout["cost_start"] + i, 2).value = field

    for offset in range(MONTH_COUNT):
        col = MONTH_START_COL + offset
        letter = ws.cell(1, col).column_letter
        for i, _field in enumerate(fields):
            crop_row = layout["crop_start"] + i
            name_row = layout["name_start"] + i
            mark_row = layout["mark_start"] + i
            labor_row = layout["labor_start"] + i
            revenue_row = layout["revenue_start"] + i
            cost_row = layout["cost_start"] + i
            ws.cell(name_row, col).value = (
                f'=IF({letter}{mark_row}="","",'
                f'VLOOKUP({letter}{crop_row},栽培スケジュール!$A:$C,3,0))'
            )
            ws.cell(labor_row, col).value = (
                f'=IF({letter}{mark_row}="","",'
                f'VLOOKUP({letter}{crop_row},栽培スケジュール!$A:$I,'
                f'IF(ISNUMBER(SEARCH("s",{letter}{mark_row})),7,'
                f'IF(ISNUMBER(SEARCH("o",{letter}{mark_row})),8,9)),0))'
            )
            ws.cell(revenue_row, col).value = (
                f'=IF(ISNUMBER(SEARCH("f",{letter}{mark_row})),'
                f'VLOOKUP({letter}{crop_row},栽培スケジュール!$A:$AA,27,0),"")'
            )
            ws.cell(cost_row, col).value = (
                f'=IF(ISNUMBER(SEARCH("s",{letter}{mark_row})),'
                f'VLOOKUP({letter}{crop_row},栽培スケジュール!$A:$AA,25,0),"")'
            )
        ws.cell(layout["labor_total"], col).value = (
            f"=SUM({letter}{layout['labor_start']}:{letter}{layout['labor_total'] - 1})"
        )
        ws.cell(layout["revenue_total"], col).value = (
            f"=SUM({letter}{layout['revenue_start']}:{letter}{layout['cost_total'] - 1})"
        )
        ws.cell(layout["cost_total"], col).value = (
            f"=SUM({letter}{layout['cost_start']}:{letter}{layout['income'] - 1})"
        )
        ws.cell(layout["income"], col).value = (
            f"={letter}{layout['revenue_total']}-{letter}{layout['cost_total']}"
        )

    set_summary_formulas(ws, layout)


def set_summary_formulas(ws, layout: dict[str, int]) -> None:
    annual_ranges = [("C", "N"), ("O", "Z"), ("AA", "AL"), ("AM", "AX")]
    year_cols = ["C", "D", "E", "F"]
    ss = layout["summary_start"]
    fill_scenario_summary(
        ws,
        ss,
        annual_ranges,
        year_cols,
        layout["revenue_total"],
        layout["cost_total"],
        add_rent=True,
        purchase=False,
    )
    fill_scenario_summary(
        ws,
        ss + 22,
        annual_ranges,
        year_cols,
        layout["revenue_total"],
        layout["cost_total"],
        add_rent=False,
        purchase=True,
    )


def fill_scenario_summary(
    ws,
    start: int,
    annual_ranges: list[tuple[str, str]],
    year_cols: list[str],
    revenue_row: int,
    cost_row: int,
    *,
    add_rent: bool,
    purchase: bool,
) -> None:
    ws.cell(start, 1).value = "農地購入の場合" if purchase else "農地賃借の場合"
    ws.cell(start + 1, 1).value = "<簡易PL>"
    ws.cell(start + 2, 2).value = "FY"
    ws.cell(start + 3, 2).value = "粗収益"
    ws.cell(start + 4, 2).value = "直接経営費"
    ws.cell(start + 5, 2).value = "減価償却費"
    ws.cell(start + 6, 2).value = "所得"
    for i, col in enumerate(year_cols):
        ws[f"{col}{start + 2}"] = i
        first, last = annual_ranges[i]
        ws[f"{col}{start + 3}"] = f"=SUM({first}{revenue_row}:{last}{revenue_row})"
        rent_add = f"+J{start + 5}" if add_rent else ""
        ws[f"{col}{start + 4}"] = f"=SUM({first}{cost_row}:{last}{cost_row}){rent_add}"
        ws[f"{col}{start + 5}"] = "=固定資産!$H$14"
        ws[f"{col}{start + 6}"] = f"={col}{start + 3}-{col}{start + 4}-{col}{start + 5}"
    if add_rent:
        ws[f"H{start + 5}"] = "土地貸借料"
        ws[f"J{start + 5}"] = "=固定資産!F21/1000"
        ws[f"K{start + 5}"] = "千円を加算"
    ws[f"H{start + 6}"] = "※平均所得"
    ws[f"J{start + 6}"] = f"=AVERAGE(D{start + 6}:F{start + 6})"
    ws[f"K{start + 6}"] = "千円"

    ws.cell(start + 8, 1).value = "<簡易PL_試行期間加味>"
    ws.cell(start + 9, 2).value = "FY"
    for i, col in enumerate(year_cols):
        ws[f"{col}{start + 9}"] = i
    ws.cell(start + 10, 2).value = "粗収益"
    ws.cell(start + 11, 2).value = "経営費"
    ws.cell(start + 12, 2).value = "減価償却費"
    ws.cell(start + 13, 2).value = "所得"
    ws[f"C{start + 10}"] = f"=C{start + 3}*0.7"
    ws[f"D{start + 10}"] = f"=D{start + 3}*0.9"
    ws[f"E{start + 10}"] = f"=E{start + 3}"
    ws[f"F{start + 10}"] = f"=F{start + 3}"
    for col in year_cols:
        ws[f"{col}{start + 11}"] = f"={col}{start + 4}"
        ws[f"{col}{start + 12}"] = f"={col}{start + 5}"
        ws[f"{col}{start + 13}"] = (
            f"={col}{start + 10}-{col}{start + 11}-{col}{start + 12}"
        )
    ws[f"H{start + 13}"] = "※平均所得"
    ws[f"J{start + 13}"] = f"=AVERAGE(D{start + 13}:F{start + 13})"
    ws[f"K{start + 13}"] = "千円"

    ws.cell(start + 15, 1).value = "<簡易CF>"
    ws.cell(start + 16, 2).value = "FY"
    ws.cell(start + 17, 2).value = "in"
    ws.cell(start + 18, 2).value = "out"
    ws.cell(start + 19, 2).value = "CF"
    ws.cell(start + 20, 2).value = "期末残"
    for i, col in enumerate(year_cols):
        ws[f"{col}{start + 16}"] = i
        ws[f"{col}{start + 17}"] = f"={col}{start + 10}"
    if purchase:
        ws[f"C{start + 18}"] = f"=C{start + 11}+固定資産!D14+I{start + 19}"
        ws[f"I{start + 19}"] = "=固定資産!F24/1000"
        ws[f"J{start + 19}"] = "千円を加算"
        initial_cash = 15000
    else:
        ws[f"C{start + 18}"] = f"=C{start + 11}+固定資産!D14"
        initial_cash = 10000
    for col in year_cols[1:]:
        ws[f"{col}{start + 18}"] = f"={col}{start + 11}*(1-$I$50)"
    for col in year_cols:
        ws[f"{col}{start + 19}"] = f"={col}{start + 17}-{col}{start + 18}"
    ws[f"C{start + 20}"] = f"=J{start + 20}+C{start + 19}"
    ws[f"D{start + 20}"] = f"=C{start + 20}+D{start + 19}"
    ws[f"E{start + 20}"] = f"=D{start + 20}+E{start + 19}"
    ws[f"F{start + 20}"] = f"=E{start + 20}+F{start + 19}"
    ws[f"H{start + 20}"] = "期初C"
    ws[f"J{start + 20}"] = initial_cash
    ws[f"K{start + 20}"] = "千円"


def excel_crop_id(crop_id: str | None) -> int | str | None:
    if crop_id is None:
        return None
    return int(crop_id) if crop_id.isdigit() else crop_id


def write_summary_sheet(
    ws,
    selected: list[Candidate],
    leasing: dict[str, list[float]],
    purchase: dict[str, list[float]],
) -> None:
    ws["A1"] = "探索結果"
    ws["A3"] = "field"
    ws["B3"] = "crop_id"
    ws["C3"] = "crop_name"
    ws["D3"] = "first_active"
    ws["E3"] = "last_harvest"
    ws["F3"] = "profit_thousand_yen"
    for row_index, candidate in enumerate(selected, start=4):
        ws.cell(row_index, 1).value = candidate.field
        ws.cell(row_index, 2).value = excel_crop_id(candidate.crop_id)
        ws.cell(row_index, 3).value = candidate.crop_name
        ws.cell(row_index, 4).value = month_label(candidate.first_active)
        ws.cell(row_index, 5).value = month_label(candidate.last_harvest)
        ws.cell(row_index, 6).value = round(candidate.profit, 3)

    metrics_start = len(selected) + 6
    ws.cell(metrics_start, 1).value = "指標"
    ws.cell(metrics_start, 2).value = "値"
    ws.cell(metrics_start + 1, 1).value = "賃借_平均所得_FY1-FY3_千円"
    ws.cell(metrics_start + 1, 2).value = sum(leasing["trial_income"][1:4]) / 3
    ws.cell(metrics_start + 2, 1).value = "購入_平均所得_FY1-FY3_千円"
    ws.cell(metrics_start + 2, 2).value = sum(purchase["trial_income"][1:4]) / 3
    ws.cell(metrics_start + 3, 1).value = "注記"
    ws.cell(metrics_start + 3, 2).value = "経営計画シートの数式はExcelで開いた時に再計算されます"


if __name__ == "__main__":
    raise SystemExit(main())
